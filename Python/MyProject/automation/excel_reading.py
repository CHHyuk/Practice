import openpyxl

wb = openpyxl.load_workbook('C:/git/practice/python/MyProject/automation/excel_cafe.xlsx')
ws = wb.active

# 시트 내 모든 데이터 출력
for x in range(1,ws.max_row): # 최대 행(row)
    for y in range(1,ws.max_column+1): # 최대 열(column)
        print(ws.cell(x,y).value, end=' ')
    print()

# cell value가 해당값일때 그 value를 아래의 밸류값으로 바꿔줌
for row in ws.iter_rows(min_row = 2): # 제목 행을 제외한 모든 대상
    for cell in row:
        if cell.value == '스타벅스':
            cell.value = '스타벅스 폐점' 

for x in range(1,ws.max_row): # 최대 행(row)
    for y in range(1,ws.max_column+1): # 최대 열(column)
        print(ws.cell(x,y).value, end=' ')
    print()
